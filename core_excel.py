# core_excel.py
import os
import warnings
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

# -------------------------- 基础配置（保留原有，抑制无关警告） --------------------------
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
warnings.filterwarnings("ignore", category=DeprecationWarning, module="openpyxl")


# -------------------------- 辅助函数：日期处理（提取年月日，剔除时分秒） --------------------------
def process_date_column(df, date_col_name):
    """
    辅助函数：处理日期列，仅保留年月日，剔除时分秒
    :param df: 待处理的DataFrame
    :param date_col_name: 日期列名称
    :return: 处理后的DataFrame
    """
    df[date_col_name] = pd.to_datetime(
        df[date_col_name],
        format=None,
        errors='coerce',
        dayfirst=False
    )
    # 提取仅日期部分（去掉时分秒），保留NaT不处理
    valid_dates = df[date_col_name].notna()
    df.loc[valid_dates, date_col_name] = df.loc[valid_dates, date_col_name].dt.date
    df[date_col_name] = df[date_col_name].fillna(pd.NaT)
    return df


# -------------------------- 辅助函数：数值列转换 --------------------------
def convert_numeric_columns(df, numeric_col_names):
    """
    辅助函数：转换指定列为数值类型
    :param df: 待处理的DataFrame
    :param numeric_col_names: 数值列名称列表
    :return: 处理后的DataFrame
    """
    for col in numeric_col_names:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')
    return df


# -------------------------- 核心Excel处理函数（脱离GUI依赖，封装为独立模块） --------------------------
def process_excel_core(template_path, data_path, save_path, log_callback=None):
    """
    核心Excel处理逻辑（与原代码功能完全一致，脱离GUI依赖）
    1.  清洗模板表原有数据（仅清空值，保留格式/公式）
    2.  提取数据文件的舟山区数据（转换数值列类型，解决公式失效问题）
    3.  提取发货表AC列（签收关单时间），仅保留日期部分，填充到模板表AS列（45列），滞留表置空
    4.  严格以模板的格式为基准，仅填充数据（不改任何格式）

    :param template_path: 模板文件路径
    :param data_path: 数据文件路径
    :param save_path: 结果文件保存路径
    :param log_callback: 日志回调函数（可选，用于输出实时日志，默认print输出）
    :return: 处理结果（布尔值）、错误信息（str，成功为None）
    """
    # 初始化日志回调函数（默认使用print输出，兼容无回调场景）
    if not log_callback:
        def default_log_callback(msg):
            timestamp = datetime.now().strftime("[%Y-%m-%d %H:%M:%S]")
            print(f"{timestamp} {msg}")

        log_callback = default_log_callback

    try:
        # 1. 验证文件是否存在
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"模板文件不存在：{template_path}")
        if not os.path.exists(data_path):
            raise FileNotFoundError(f"数据文件不存在：{data_path}")

        log_callback("开始读取数据文件（提取数据并转换数值类型）...")

        # 2. 读取数据文件，模糊匹配发货/滞留表
        xl = pd.ExcelFile(data_path)
        ship_sheet_name = [name for name in xl.sheet_names if "发货" in name][0]
        delay_sheet_name = [name for name in xl.sheet_names if "滞留" in name][0]

        log_callback(f"识别到数据文件发货表：{ship_sheet_name}，滞留表：{delay_sheet_name}")

        # 3. 处理发货表数据（提取指定列、转换数值类型、处理日期、筛选舟山区）
        log_callback("提取数据文件发货表舟山区数据并转换数值类型...")
        df_ship = pd.read_excel(data_path, sheet_name=ship_sheet_name)
        # 原代码指定列索引（保持不变，确保功能一致）
        ship_cols = {
            '城市群': df_ship.iloc[:, 3],
            '客户名称': df_ship.iloc[:, 7],
            '物品说明': df_ship.iloc[:, 15],
            '发货数量': df_ship.iloc[:, 21],
            '净重(吨)': df_ship.iloc[:, 24],
            '价目表价格': df_ship.iloc[:, 26],
            '税率': df_ship.iloc[:, 43],
            '签收关单时间': df_ship.iloc[:, 28]  # AC列：签收关单时间（带时分秒）
        }
        df_ship_processed = pd.DataFrame(ship_cols)

        # 转换数值型列为数字类型（复用辅助函数）
        numeric_cols = ['发货数量', '净重(吨)', '价目表价格', '税率']
        df_ship_processed = convert_numeric_columns(df_ship_processed, numeric_cols)

        # 处理日期列：仅保留年月日，剔除时分秒（复用辅助函数）
        log_callback("转换发货表签收关单时间（仅保留年月日，隐藏时分秒）...")
        df_ship_processed = process_date_column(df_ship_processed, '签收关单时间')

        # 筛选舟山区数据（保持原代码硬编码，后续可优化为配置项）
        df_ship_zhoushan = df_ship_processed[df_ship_processed['城市群'] == '舟山区'].copy()
        log_callback(f"数据文件发货表提取到舟山区数据 {len(df_ship_zhoushan)} 行")

        # 4. 处理滞留表数据（提取指定列、转换数值类型、置空日期、筛选舟山区有效数据）
        log_callback("提取数据文件滞留表舟山区数据并转换数值类型...")
        df_delay = pd.read_excel(data_path, sheet_name=delay_sheet_name)
        # 原代码指定列索引（保持不变，确保功能一致）
        delay_cols = {
            '城市群': df_delay.iloc[:, 3],
            '客户名称': df_delay.iloc[:, 5],
            '物品说明': df_delay.iloc[:, 18],
            '发货数量': df_delay.iloc[:, 27],  # 订货数量对应发货数量
            '净重(吨)': df_delay.iloc[:, 30],
            '价目表价格': df_delay.iloc[:, 33],
            '税率': df_delay.iloc[:, 55],
            '发运库存组织': df_delay.iloc[:, 26],  # 用于剔除空值
            '签收关单时间': pd.Series(dtype='object')  # 初始化空列，兼容日期格式
        }
        df_delay_processed = pd.DataFrame(delay_cols)

        # 转换滞留表数值列类型
        df_delay_processed = convert_numeric_columns(df_delay_processed, numeric_cols)

        # 滞留表签收关单时间统一置空（用None，兼容openpyxl）
        df_delay_processed['签收关单时间'] = None

        # 筛选舟山区 + 剔除发运库存组织为空的行
        df_delay_zhoushan = df_delay_processed[
            (df_delay_processed['城市群'] == '舟山区') &
            (df_delay_processed['发运库存组织'].notna()) &
            (df_delay_processed['发运库存组织'] != '')
            ].copy()
        df_delay_zhoushan = df_delay_zhoushan.drop('发运库存组织', axis=1)
        log_callback(f"数据文件滞留表提取到舟山区数据 {len(df_delay_zhoushan)} 行")

        # 5. 合并发货/滞留数据，剔除全空行
        log_callback("合并数据文件发货/滞留数据（仅保留日期，隐藏时分秒）...")
        df_combined = pd.concat([df_ship_zhoushan, df_delay_zhoushan], ignore_index=True)
        df_combined = df_combined.dropna(how='all')
        log_callback(f"数据文件合并后总数据 {len(df_combined)} 行")

        if len(df_combined) == 0:
            log_callback("警告：数据文件合并后无有效数据！")
            return False, "数据文件合并后无有效数据"

        # 6. 加载模板文件（保留所有原始格式/公式）
        log_callback("加载模板文件（保留所有原始格式）...")
        wb = load_workbook(template_path, data_only=False, keep_links=False)
        if "宁波发货" not in wb.sheetnames:
            raise KeyError("模板文件中不存在'宁波发货'工作表")
        ws = wb["宁波发货"]

        # 7. 数据清洗：清空模板表原有数据（仅清空值，保留格式/公式，AS列=45）
        log_callback("清洗模板表原有数据（仅清空数值，保留格式和公式）...")
        clean_columns = [4, 8, 15, 21, 24, 26, 43, 45]  # AS列=45（修正后）
        max_row = ws.max_row
        # 从第2行开始清空（第1行是表头，保留不修改）
        for row in range(2, max_row + 1):
            for col in clean_columns:
                ws.cell(row=row, column=col).value = None  # 仅清空单元格值，保留格式

        # 8. 填充新数据到模板（保留原格式，AS列=45设置日期格式）
        log_callback("填充数据文件新数据到模板（仅保留日期，兼容Excel公式）...")
        start_row = 2  # 从第2行开始（避开表头）
        for idx, (_, row) in enumerate(df_combined.iterrows()):
            current_row = start_row + idx
            # 文本列赋值（保持原列索引，确保格式匹配）
            ws.cell(row=current_row, column=4).value = row['城市群']  # D列 城市群
            ws.cell(row=current_row, column=8).value = row['客户名称']  # H列 客户名称
            ws.cell(row=current_row, column=15).value = row['物品说明']  # O列 物品说明

            # 数值列赋值+设置数字格式（保持原格式，确保Excel公式兼容）
            ws.cell(row=current_row, column=21).value = row['发货数量']
            ws.cell(row=current_row, column=21).number_format = '0.00'

            ws.cell(row=current_row, column=24).value = row['净重(吨)']
            ws.cell(row=current_row, column=24).number_format = '0.000'

            ws.cell(row=current_row, column=26).value = row['价目表价格']
            ws.cell(row=current_row, column=26).number_format = '0.00'

            ws.cell(row=current_row, column=43).value = row['税率']
            ws.cell(row=current_row, column=43).number_format = '0.00%'

            # AS列=45 日期赋值+格式设置（仅保留年月日）
            ws.cell(row=current_row, column=45).value = row['签收关单时间']
            ws.cell(row=current_row, column=45).number_format = 'yyyy/mm/dd'

        # 9. 保存结果文件（保留模板格式，兼容Excel公式）
        wb.save(save_path)
        wb.close()

        log_callback(f"处理完成！模板表原有数据已清洗，AS列（45列）仅保留日期部分，保存至：{save_path}")
        return True, None

    except Exception as e:
        error_msg = f"处理失败：{str(e)}"
        log_callback(error_msg)
        return False, error_msg